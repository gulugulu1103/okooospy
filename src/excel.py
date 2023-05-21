import logging
import os
from typing import List

import win32com.client
from openpyxl import *
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import classes


def set_header_style(cell):
	cell.font = Font(bold = True, size = 14)
	cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
	cell.fill = PatternFill(start_color = 'D9D9D9', end_color = 'D9D9D9', fill_type = 'solid')
	cell.border = Border(left = Side(style = 'thin'), right = Side(style = 'thin'), top = Side(style = 'thin'),
	                     bottom = Side(style = 'double'))


def write_match_list(data: list):
	if os.path.exists('output.xlsx'):
		workbook = load_workbook(filename = 'output.xlsx', data_only = False)
	else:
		workbook = Workbook()

	# 获取工作簿中的所有工作表
	worksheets = workbook.worksheets

	# 检查第一个工作表是否命名为“场次列表”
	if worksheets[0].title == '场次列表':
		match_list_worksheet = worksheets[0]
	else:
		# 创建一个场次列表
		match_list_worksheet = workbook.create_sheet('场次列表', 0)

	# 写入表头
	headers = ['场次', '赛事', '比赛时间', '主', '比分', '客', 'ID', '选中']
	for col_num, header in enumerate(headers, 1):
		match_list_worksheet.cell(row = 1, column = col_num).value = header
		set_header_style(match_list_worksheet.cell(row = 1, column = col_num))

	# 设置列宽
	column_widths = [5, 10, 20, 20, 10, 20, 10]
	for col_num, width in enumerate(column_widths, 1):
		match_list_worksheet.column_dimensions[
			match_list_worksheet.cell(row = 1, column = col_num).column_letter].width = width

	# 将查询结果写入工作表
	for row_num, row in enumerate(data, 2):
		match_list_worksheet.cell(row = row_num, column = 1).value = row_num - 1  # 场次
		match_list_worksheet.cell(row = row_num, column = 2).value = row[0]  # 赛事
		match_list_worksheet.cell(row = row_num, column = 3).value = f"{row[1]} {row[2]}"  # 比赛时间
		match_list_worksheet.cell(row = row_num, column = 4).value = row[3]  # 主队
		if row[6] and row[7]:
			match_list_worksheet.cell(row = row_num, column = 5).value = f"{row[6]} - {row[7]}"  # 比分
		match_list_worksheet.cell(row = row_num, column = 6).value = row[4]  # 客队
		match_list_worksheet.cell(row = row_num, column = 7).value = row[5]  # ID

	workbook.save('output.xlsx')
	workbook.close()


def get_selected_ids(workbook_path="output.xlsx", sheet_name="场次列表"):
	# 加载工作簿和工作表
	wb = load_workbook(workbook_path, data_only = True)
	match_list_worksheet = wb[sheet_name]

	selected_ids = []

	# 遍历“选中”列
	for row_num in range(2, match_list_worksheet.max_row + 1):
		selected = match_list_worksheet.cell(row = row_num, column = 8).value
		if selected == 1:
			# 获取相应的 ID 并添加到列表中
			match_id = match_list_worksheet.cell(row = row_num, column = 7).value
			selected_ids.append(match_id)

	return selected_ids


def fill_area(worksheet, row_range, column_range, match_data, data_key):
	for i, list in zip(row_range, match_data.values()):
		if data_key in list:
			for j, value in zip(column_range, list[data_key]):
				worksheet.cell(row = i, column = j, value = value)


def merge_workbooks(filename='../output.xlsm'):
	output_file_path = os.path.abspath(filename)

	# 执行CopySheet函数，该函数会将模板中的模板复制为比赛ID的Sheet
	xls = win32com.client.Dispatch("Excel.Application")
	wb = xls.Workbooks.Open(output_file_path)
	xls.Run("MergeSheets")
	wb.Save()
	wb.Close()  # 关闭工作簿
	xls.Quit()  # 退出 Excel 应用


def fill_in_workbook(match_list, filename='../output.xlsm'):
	# 加载工作簿和工作表
	workbook = load_workbook(filename = filename, data_only = False, keep_vba = True)

	# 迭代比赛列表
	for match in match_list:
		worksheet = workbook[str(match.match_id)]
		# 比赛基本信息
		worksheet.cell(row = 1, column = 1, value = f"{match.league} {match.match_time.split()[-1]}")
		worksheet.cell(row = 1, column = 2, value = match.home_team)
		worksheet.cell(row = 1, column = 6, value = match.away_team)

		# 开始遍历工作表中的A列
		logging.info("开始遍历工作表中的公司名称")
		for row in range(2, worksheet.max_row + 1):
			# 读取A列的公司名称
			company = worksheet.cell(row = row, column = 1).value
			# 如果公司名称在 match.odds 字典中，则获取其对应的数据，否则赋值为 None
			odds_info = match.odds.get(company) if company in match.odds else None
			# 如果公司的赔率信息存在
			if odds_info:
				logging.info(f"在 {company} 的赔率信息中填入数据")
				# 判断 'initial' 字段是否存在，并将其数据填入到 B、C、D 列
				if 'initial' in odds_info:
					worksheet.cell(row = row, column = 2).value = odds_info['initial'][0]
					worksheet.cell(row = row, column = 3).value = odds_info['initial'][1]
					worksheet.cell(row = row, column = 4).value = odds_info['initial'][2]
				# 判断 'current' 字段是否存在，并将其数据填入到 F、G、H 列
				if 'current' in odds_info:
					worksheet.cell(row = row, column = 6).value = odds_info['current'][0]
					worksheet.cell(row = row, column = 7).value = odds_info['current'][1]
					worksheet.cell(row = row, column = 8).value = odds_info['current'][2]
				# 判断 'updated' 字段是否存在，并将其数据填入到 J、K、L 列
				if 'updated' in odds_info:
					worksheet.cell(row = row, column = 10).value = odds_info['updated'][0]
					worksheet.cell(row = row, column = 11).value = odds_info['updated'][1]
					worksheet.cell(row = row, column = 12).value = odds_info['updated'][2]
				# 判断 'current_kelly' 字段是否存在，并将其数据填入到 N、O、P 列
				if 'current_kelly' in odds_info:
					worksheet.cell(row = row, column = 14).value = odds_info['current_kelly'][0]
					worksheet.cell(row = row, column = 15).value = odds_info['current_kelly'][1]
					worksheet.cell(row = row, column = 16).value = odds_info['current_kelly'][2]
				# 判断 'updated_kelly' 字段是否存在，并将其数据填入到 Q、R、S 列
				if 'updated_kelly' in odds_info:
					worksheet.cell(row = row, column = 17).value = odds_info['updated_kelly'][0]
					worksheet.cell(row = row, column = 18).value = odds_info['updated_kelly'][1]
					worksheet.cell(row = row, column = 19).value = odds_info['updated_kelly'][2]

		# 所有数据填入完成，保存工作簿到文件
		logging.info("赔率和凯利信息填入完成")

		# 填入亚盘信息
		logging.info("开始填入亚盘信息")
		# 定义需要填充数据的公司列表
		companies = ['澳门彩票', 'Interwetten', '金宝博']
		# 按照'initial'、'current' 和 'updated' 的顺序，对每个公司进行操作
		for i, company in enumerate(companies):
			logging.info(f"开始处理 {company}的亚盘信息")
			if company in match.asian_handicaps:
				asian_handicaps_info = match.asian_handicaps.get(company)
				if asian_handicaps_info:
					# 填入 'initial' 数据
					if 'initial' in asian_handicaps_info:
						worksheet.cell(row = i + 2, column = 23, value = asian_handicaps_info['initial'][0])
						worksheet.cell(row = i + 2, column = 24, value = asian_handicaps_info['initial'][1])
						worksheet.cell(row = i + 2, column = 25, value = asian_handicaps_info['initial'][2])
					# 填入 'current' 数据
					if 'current' in asian_handicaps_info:
						worksheet.cell(row = i + 5, column = 23, value = asian_handicaps_info['current'][0])
						worksheet.cell(row = i + 5, column = 24, value = asian_handicaps_info['current'][1])
						worksheet.cell(row = i + 5, column = 25, value = asian_handicaps_info['current'][2])
					# 填入 'updated' 数据
					if 'updated' in asian_handicaps_info:
						worksheet.cell(row = i + 8, column = 23, value = asian_handicaps_info['updated'][0])
						worksheet.cell(row = i + 8, column = 24, value = asian_handicaps_info['updated'][1])
						worksheet.cell(row = i + 8, column = 25, value = asian_handicaps_info['updated'][2])
			else:
				logging.info(f"{company} 的亚盘信息不存在")

		# 比赛让球值
		worksheet.cell(row = 1, column = 27, value = match.handicap)

		# 填入让球盘信息
		logging.info("开始填入让球盘信息")
		# 定义需要填充数据的公司列表
		companies = ['竞彩官方', 'Interwetten', '金宝博']

		# 按照公司列表的顺序，对每个公司进行操作
		for i, company in enumerate(companies):
			logging.info(f"开始处理 {company}")
			if company in match.handicap_odds:
				handicap_odds_info = match.handicap_odds.get(company)
				if handicap_odds_info:
					# 填入 'initial' 数据
					if 'initial' in handicap_odds_info:
						worksheet.cell(row = i + 2, column = 28, value = handicap_odds_info['initial'][0])
						worksheet.cell(row = i + 2, column = 29, value = handicap_odds_info['initial'][1])
						worksheet.cell(row = i + 2, column = 30, value = handicap_odds_info['initial'][2])
					# 填入 'updated' 数据, 如果没有 'updated' 数据则填入 'current' 数据
					if 'updated' in handicap_odds_info:
						worksheet.cell(row = i + 5, column = 28, value = handicap_odds_info['updated'][0])
						worksheet.cell(row = i + 5, column = 29, value = handicap_odds_info['updated'][1])
						worksheet.cell(row = i + 5, column = 30, value = handicap_odds_info['updated'][2])
					elif 'current' in handicap_odds_info:
						worksheet.cell(row = i + 5, column = 28, value = handicap_odds_info['current'][0])
						worksheet.cell(row = i + 5, column = 29, value = handicap_odds_info['current'][1])
						worksheet.cell(row = i + 5, column = 30, value = handicap_odds_info['current'][2])
				else:
					logging.info(f"{company} 的让球指数信息不存在")
			else:
				logging.info(f"{company} 的让球指数信息不存在")

		# 填入盈亏信息
		logging.info("开始填入盈亏信息")
		# 定义需要填充数据的公司列表
		companies = ['必发', '竞彩']
		# 按照'initial' 和 'updated' 的顺序，对每个公司进行操作
		for i, company in enumerate(companies):
			logging.info(f"开始处理 {company}")
			if company in match.profit_loss:
				profit_loss_info = match.profit_loss.get(company)
				if profit_loss_info:
					# 填入 'initial' 数据
					if 'initial' in profit_loss_info:
						worksheet.cell(row = i + 9, column = 28, value = profit_loss_info['initial'][0])
						worksheet.cell(row = i + 9, column = 29, value = profit_loss_info['initial'][1])
						worksheet.cell(row = i + 9, column = 30, value = profit_loss_info['initial'][2])
					# 填入 'updated' 数据
					if 'updated' in profit_loss_info:
						worksheet.cell(row = i + 11, column = 28, value = profit_loss_info['updated'][0])
						worksheet.cell(row = i + 11, column = 29, value = profit_loss_info['updated'][1])
						worksheet.cell(row = i + 11, column = 30, value = profit_loss_info['updated'][2])
			else:
				logging.info(f"{company} 的盈亏信息不存在")

		# 填入比赛历史信息
		logging.info("开始填入比赛历史信息")
		# 定义需要填充数据的键列表
		keys = ['主队历史', '客队历史', '交战历史', '最后交战历史']
		# 如果match.match_history存在，进行数据填充
		if match.match_history:
			# 按照键列表的顺序，对每个键进行操作
			for i, key in zip(range(2, 11, 2), keys):
				logging.info(f"开始处理 {key}")
				if key in match.match_history:
					worksheet.cell(row = i, column = 32, value = match.match_history[key])
				else:
					logging.info(f"{key} 的历史信息不存在")
		else:
			logging.info("比赛历史信息不存在")

	workbook.save(filename = filename)


def prepare_workbook(match_list,
                     template_filename: str = '../template.xlsm',
                     output_filename: str = '../output.xlsm'):
	# 读取模板，复制为output.xlsm
	import shutil
	output_file_path = os.path.abspath(output_filename)
	shutil.copy2(template_filename, output_filename)

	# 执行CopySheet函数，该函数会将模板中的模板复制为比赛ID的Sheet
	xls = win32com.client.Dispatch("Excel.Application")
	wb = xls.Workbooks.Open(output_file_path)
	for match in match_list:
		xls.Run("CopySheet", str(match.match_id))
	xls.Run("DeleteTemplate")
	wb.Save()
	wb.Close()  # 关闭工作簿
	xls.Quit()  # 退出 Excel 应用


def generate_merged_xlsm(matchlist: List[classes.SoccerMatch], template_filename='../template.xlsm',
                         output_filename='../output.xlsm', open_file: bool = True):
	"""
	生成合并后的Excel文件
	Args:
		matchlist:
		template_filename:
		output_filename:

	Returns:
	"""
	template_file_path = os.path.abspath(template_filename)
	output_file_path = os.path.abspath(output_filename)
	prepare_workbook(matchlist, template_filename = template_file_path, output_filename = output_file_path)
	fill_in_workbook(matchlist, filename = output_file_path)
	merge_workbooks(filename = output_file_path)
	if open_file:
		os.startfile(output_file_path)


if __name__ == "__main__":
	pass
