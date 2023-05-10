import copy
import os

import openpyxl.worksheet.worksheet
from openpyxl import *
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import test
from classes import *

def set_header_style(cell):
	cell.font = Font(bold = True, size = 14)
	cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
	cell.fill = PatternFill(start_color = 'D9D9D9', end_color = 'D9D9D9', fill_type = 'solid')
	cell.border = Border(left = Side(style = 'thin'), right = Side(style = 'thin'), top = Side(style = 'thin'),
	                     bottom = Side(style = 'double'))


def write_match_list(data: list):
	if os.path.exists('output.xlsx'):
		workbook = load_workbook(filename = 'output.xlsx', data_only = False)
	else:
		workbook = Workbook()

	# 获取工作簿中的所有工作表
	worksheets = workbook.worksheets

	# 检查第一个工作表是否命名为“场次列表”
	if worksheets[0].title == '场次列表':
		match_list_worksheet = worksheets[0]
	else:
		# 创建一个场次列表
		match_list_worksheet = workbook.create_sheet('场次列表', 0)

	# 写入表头
	headers = ['场次', '赛事', '比赛时间', '主', '比分', '客', 'ID', '选中']
	for col_num, header in enumerate(headers, 1):
		match_list_worksheet.cell(row = 1, column = col_num).value = header
		set_header_style(match_list_worksheet.cell(row = 1, column = col_num))

	# 设置列宽
	column_widths = [5, 10, 20, 20, 10, 20, 10]
	for col_num, width in enumerate(column_widths, 1):
		match_list_worksheet.column_dimensions[
			match_list_worksheet.cell(row = 1, column = col_num).column_letter].width = width

	# 将查询结果写入工作表
	for row_num, row in enumerate(data, 2):
		match_list_worksheet.cell(row = row_num, column = 1).value = row_num - 1  # 场次
		match_list_worksheet.cell(row = row_num, column = 2).value = row[0]  # 赛事
		match_list_worksheet.cell(row = row_num, column = 3).value = f"{row[1]} {row[2]}"  # 比赛时间
		match_list_worksheet.cell(row = row_num, column = 4).value = row[3]  # 主队
		if row[6] and row[7]:
			match_list_worksheet.cell(row = row_num, column = 5).value = f"{row[6]} - {row[7]}"  # 比分
		match_list_worksheet.cell(row = row_num, column = 6).value = row[4]  # 客队
		match_list_worksheet.cell(row = row_num, column = 7).value = row[5]  # ID

	workbook.save('output.xlsx')
	workbook.close()


def get_selected_ids(workbook_path="output.xlsx", sheet_name="场次列表"):
	# 加载工作簿和工作表
	wb = load_workbook(workbook_path, data_only = True)
	match_list_worksheet = wb[sheet_name]

	selected_ids = []

	# 遍历“选中”列
	for row_num in range(2, match_list_worksheet.max_row + 1):
		selected = match_list_worksheet.cell(row = row_num, column = 8).value
		if selected == 1:
			# 获取相应的 ID 并添加到列表中
			match_id = match_list_worksheet.cell(row = row_num, column = 7).value
			selected_ids.append(match_id)

	return selected_ids


def fill_area(worksheet, row_range, column_range, match_data, data_key):
	for i, list in zip(row_range, match_data.values()):
		if data_key in list:
			for j, value in zip(column_range, list[data_key]):
				worksheet.cell(row = i, column = j, value = value)


def copy_cell(src_cell, dst_cell):
	dst_cell.value = src_cell.value
	dst_cell.data_type = src_cell.data_type
	dst_cell.font = copy.copy(src_cell.font)
	dst_cell.border = copy.copy(src_cell.border)
	dst_cell.fill = copy.copy(src_cell.fill)
	dst_cell.number_format = src_cell.number_format
	dst_cell.protection = copy.copy(src_cell.protection)
	dst_cell.alignment = copy.copy(src_cell.alignment)


def merge_worksheets(template_file: str, data_list):
	merged_wb = openpyxl.Workbook()
	merged_sheet = merged_wb.active
	merged_sheet.title = "Merged"

	row_offset = 0
	for data in data_list:
		template_wb = openpyxl.load_workbook(template_file, data_only = False)
		fill_template(template_wb, data)
		template_sheet = template_wb.active

		max_row = template_sheet.max_row
		max_col = template_sheet.max_column

		for row in range(1, max_row + 1):
			for col in range(1, max_col + 1):
				src_cell = template_sheet.cell(row = row, column = col)
				dst_cell = merged_sheet.cell(row = row + row_offset, column = col)
				copy_cell(src_cell, dst_cell)

		row_offset += max_row + 1

	return merged_wb


def fill_template(template_wb, match):
	worksheet = template_wb.active

	# 比赛基本信息
	worksheet.cell(row = 1, column = 1, value = f"{match.league} {match.match_time.split()[-1]}")
	worksheet.cell(row = 1, column = 2, value = match.home_team)
	worksheet.cell(row = 1, column = 6, value = match.away_team)

	# 比赛初始赔率
	fill_area(worksheet, range(2, 13), range(2, 5), match.odds, 'initial')
	# 比赛即时赔率
	fill_area(worksheet, range(2, 13), range(6, 9), match.odds, 'current')
	# 比赛更新赔率
	fill_area(worksheet, range(2, 13), range(10, 13), match.odds, 'updated')
	# 比赛即时凯利指数
	fill_area(worksheet, range(2, 13), range(14, 17), match.odds, 'current_kelly')
	# 比赛更新凯利指数
	fill_area(worksheet, range(2, 13), range(17, 20), match.odds, 'updated_kelly')

	# 比赛初始亚盘
	fill_area(worksheet, range(2, 5), range(23, 26), match.asian_handicaps, 'initial')
	# 比赛即时亚盘
	fill_area(worksheet, range(5, 8), range(23, 26), match.asian_handicaps, 'current')
	# 比赛更新亚盘
	fill_area(worksheet, range(8, 11), range(23, 26), match.asian_handicaps, 'updated')

	# 比赛让球值
	worksheet.cell(row = 1, column = 27, value = match.handicap)
	# 比赛各个平台让球值
	fill_area(worksheet, range(2, 5), range(28, 31), match.handicap_odds, 'initial')
	# 比赛各个平台更新让球值
	fill_area(worksheet, range(5, 8), range(28, 31), match.handicap_odds, 'updated')
	# 比赛盈亏值
	fill_area(worksheet, range(9, 11), range(28, 31), match.profit_loss, 'initial')
	# 比赛更新盈亏值
	fill_area(worksheet, range(11, 13), range(28, 31), match.profit_loss, 'updated')

	# 比赛历史
	for i, string in zip(range(2, 11, 2), match.match_history.values()):
		worksheet.cell(row = i, column = 32, value = string)


if __name__ == "__main__":
	match = test.read_update_sample_match()
	merged_wb = merge_worksheets('template.xlsx', [match, match, match])
	merged_wb.save('merged_workbook.xlsx')
